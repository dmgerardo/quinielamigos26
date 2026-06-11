# -*- coding: utf-8 -*-
"""Genera js/tournament-data.js a partir de worldcup-soccer-2026-2.xlsx"""
import openpyxl, datetime as dt, json, shutil, tempfile, os

SRC = 'worldcup-soccer-2026-2.xlsx'
try:
    wb = openpyxl.load_workbook(SRC, data_only=True)
except PermissionError:
    # OneDrive/Excel a veces bloquea el archivo: leemos una copia temporal
    tmp = os.path.join(tempfile.gettempdir(), '_wc_fixtures_tmp.xlsx')
    shutil.copyfile(SRC, tmp)
    wb = openpyxl.load_workbook(tmp, data_only=True)
gs = wb['Groups']

# nombre EN -> (es, bandera)
NAME = {
 'Mexico':('M├йxico','ЁЯЗ▓ЁЯЗ╜'),'South Africa':('Sud├бfrica','ЁЯЗ┐ЁЯЗж'),'South Korea':('Corea del Sur','ЁЯЗ░ЁЯЗ╖'),
 'Czechia':('Chequia','ЁЯЗиЁЯЗ┐'),'Canada':('Canad├б','ЁЯЗиЁЯЗж'),'Bosnia-Herzegovina':('Bosnia y Herzegovina','ЁЯЗзЁЯЗж'),
 'Qatar':('Catar','ЁЯЗ╢ЁЯЗж'),'Switzerland':('Suiza','ЁЯЗиЁЯЗн'),'Brazil':('Brasil','ЁЯЗзЁЯЗ╖'),'Morocco':('Marruecos','ЁЯЗ▓ЁЯЗж'),
 'Haiti':('Hait├н','ЁЯЗнЁЯЗ╣'),'Scotland':('Escocia','ЁЯП┤ґаБзґаБвґаБ│ґаБгґаБ┤ґаБ┐'),'USA':('Estados Unidos','ЁЯЗ║ЁЯЗ╕'),'Paraguay':('Paraguay','ЁЯЗ╡ЁЯЗ╛'),
 'Australia':('Australia','ЁЯЗжЁЯЗ║'),'Germany':('Alemania','ЁЯЗйЁЯЗк'),'Curacao':('Curazao','ЁЯЗиЁЯЗ╝'),'Ecuador':('Ecuador','ЁЯЗкЁЯЗи'),
 'Netherlands':('Pa├нses Bajos','ЁЯЗ│ЁЯЗ▒'),'Japan':('Jap├│n','ЁЯЗпЁЯЗ╡'),'Sweden':('Suecia','ЁЯЗ╕ЁЯЗк'),'Tunisia':('T├║nez','ЁЯЗ╣ЁЯЗ│'),
 'Belgium':('B├йlgica','ЁЯЗзЁЯЗк'),'Egypt':('Egipto','ЁЯЗкЁЯЗм'),'IR Iran':('Ir├бn','ЁЯЗоЁЯЗ╖'),'New Zealand':('Nueva Zelanda','ЁЯЗ│ЁЯЗ┐'),
 'Spain':('Espa├▒a','ЁЯЗкЁЯЗ╕'),'Cabo Verde':('Cabo Verde','ЁЯЗиЁЯЗ╗'),'Saudi Arabia':('Arabia Saudita','ЁЯЗ╕ЁЯЗж'),
 'Uruguay':('Uruguay','ЁЯЗ║ЁЯЗ╛'),'France':('Francia','ЁЯЗлЁЯЗ╖'),'Senegal':('Senegal','ЁЯЗ╕ЁЯЗ│'),'Iraq':('Irak','ЁЯЗоЁЯЗ╢'),
 'Norway':('Noruega','ЁЯЗ│ЁЯЗ┤'),'Argentina':('Argentina','ЁЯЗжЁЯЗ╖'),'Algeria':('Argelia','ЁЯЗйЁЯЗ┐'),'Austria':('Austria','ЁЯЗжЁЯЗ╣'),
 'Jordan':('Jordania','ЁЯЗпЁЯЗ┤'),'Portugal':('Portugal','ЁЯЗ╡ЁЯЗ╣'),'Congo DR':('RD Congo','ЁЯЗиЁЯЗй'),
 'Uzbekistan':('Uzbekist├бn','ЁЯЗ║ЁЯЗ┐'),'Colombia':('Colombia','ЁЯЗиЁЯЗ┤'),'England':('Inglaterra','ЁЯП┤ґаБзґаБвґаБеґаБоґаБзґаБ┐'),
 'Croatia':('Croacia','ЁЯЗнЁЯЗ╖'),'Ghana':('Ghana','ЁЯЗмЁЯЗн'),'Panama':('Panam├б','ЁЯЗ╡ЁЯЗж'),
}
def resolve(name):
    n = str(name).strip()
    if 'Ivoire' in n: return ('Costa de Marfil','ЁЯЗиЁЯЗо')
    if 'rkiye' in n or 'Turk' in n: return ('Turqu├нa','ЁЯЗ╣ЁЯЗ╖')
    return NAME[n]

WD=['Lun','Mar','Mi├й','Jue','Vie','S├бb','Dom']
MON=['ene','feb','mar','abr','may','jun','jul','ago','sep','oct','nov','dic']
def pdate(v):
    if isinstance(v,(dt.datetime,)): return v.date()
    if isinstance(v,dt.date): return v
    s=str(v).strip()
    if ',' in s: s=s.split(',')[1].strip()
    m,d,y=s.split('.'); return dt.date(int(y),int(m),int(d))
def ptime(v):
    if isinstance(v,dt.time): return v.hour, v.minute
    p=str(v).strip().split(':'); return int(p[0]), int(p[1])
# Los horarios del fixture est├бn en hora del Este (EDT = UTC-4 en jun-jul 2026).
# Guardamos un instante absoluto en UTC; el navegador lo muestra en la zona del usuario.
def et_to_utc(d, hh, mm):
    # interpreta hh:mm como hora ET (UTC-4) y devuelve el instante en UTC
    return dt.datetime(d.year, d.month, d.day, hh, mm, tzinfo=dt.timezone.utc) + dt.timedelta(hours=4)

HOME=[2,5,8,11,14,17]  # columnas de cada grupo en el bloque
# (letra_grupo, fila_equipos_top, filas_de_jornada(city,date,teams)x6)
BLOCKS=[
 (['A','B','C','D','E','F'],4,[(9,10,11),(14,15,16),(19,20,21),(24,25,26),(29,30,31),(34,35,36)]),
 (['G','H','I','J','K','L'],46,[(51,52,53),(56,57,58),(61,62,63),(66,67,68),(71,72,73),(76,77,78)]),
]

groups={}   # letra -> [(es,flag)x4]
gmatches=[] # (grupo, home(es,flag), away(es,flag), date_str, venue)
for letters,toprow,days in BLOCKS:
    for gi,letter in enumerate(letters):
        col=HOME[gi]
        groups[letter]=[resolve(gs.cell(toprow+k,col).value) for k in range(4)]
        for (crow,drow,trow) in days:
            venue=gs.cell(crow,col).value
            d=pdate(gs.cell(drow,col).value); hh,mm=ptime(gs.cell(drow,col+1).value)
            h=resolve(gs.cell(trow,col).value); a=resolve(gs.cell(trow,col+1).value)
            gmatches.append((letter,h,a,et_to_utc(d,hh,mm),str(venue).strip()))

# --- Eliminatorias (num, slotA, slotB, sede, fecha ISO, hora) ---
def slot(s):
    s=str(s).strip()
    if s=='*': return 'Mejor 3┬║'
    if s.startswith('1. Gr.'): return '1┬║ Grupo '+s.split('.')[-1].strip()
    if s.startswith('2. Gr.'): return '2┬║ Grupo '+s.split('.')[-1].strip()
    if s.startswith('Winner'): return 'Ganador #'+s.split()[-1]
    if s.startswith('Loser') or s.startswith('Verlierer'): return 'Perdedor #'+s.split()[-1]
    return s

KO=[
 ('dieciseisavos',[
  (73,'2. Gr. A','2. Gr. B','Los Angeles','2026-06-28','15:00'),
  (74,'1. Gr. E','*','Boston','2026-06-29','16:30'),
  (75,'1. Gr. F','2. Gr. C','Monterrey','2026-06-29','21:00'),
  (76,'1. Gr. C','2. Gr. F','Houston','2026-06-29','13:00'),
  (77,'1. Gr. I','*','New York / New Jersey','2026-06-30','17:00'),
  (78,'2. Gr. E','2. Gr. I','Dallas','2026-06-30','13:00'),
  (79,'1. Gr. A','*','Mexico City','2026-06-30','21:00'),
  (80,'1. Gr. L','*','Atlanta','2026-07-01','12:00'),
  (81,'1. Gr. D','*','San Francisco Bay Area','2026-07-01','20:00'),
  (82,'1. Gr. G','*','Seattle','2026-07-01','16:00'),
  (83,'2. Gr. K','2. Gr. L','Toronto','2026-07-02','19:00'),
  (84,'1. Gr. H','2. Gr. J','Los Angeles','2026-07-02','15:00'),
  (85,'1. Gr. B','*','Vancouver','2026-07-02','23:00'),
  (86,'1. Gr. J','2. Gr. H','Miami','2026-07-04','18:00'),
  (87,'1. Gr. K','*','Kansas City','2026-07-03','21:30'),
  (88,'2. Gr. D','2. Gr. G','Dallas','2026-07-03','14:00'),
 ]),
 ('octavos',[
  (89,'Winner 74','Winner 77','Philadelphia','2026-07-04','17:00'),
  (90,'Winner 73','Winner 75','Houston','2026-07-04','13:00'),
  (91,'Winner 76','Winner 78','New York / New Jersey','2026-07-05','16:00'),
  (92,'Winner 79','Winner 80','Mexico City','2026-07-05','20:00'),
  (93,'Winner 83','Winner 84','Dallas','2026-07-06','15:00'),
  (94,'Winner 81','Winner 82','Seattle','2026-07-06','20:00'),
  (95,'Winner 86','Winner 88','Atlanta','2026-07-07','12:00'),
  (96,'Winner 85','Winner 87','Vancouver','2026-07-07','16:00'),
 ]),
 ('cuartos',[
  (97,'Winner 89','Winner 90','Boston','2026-07-09','16:00'),
  (98,'Winner 93','Winner 94','Los Angeles','2026-07-10','15:00'),
  (99,'Winner 91','Winner 92','Miami','2026-07-11','17:00'),
  (100,'Winner 95','Winner 96','Kansas City','2026-07-11','21:00'),
 ]),
 ('semis',[
  (101,'Winner 97','Winner 98','Dallas','2026-07-14','15:00'),
  (102,'Winner 99','Winner 100','Atlanta','2026-07-15','15:00'),
 ]),
 ('tercer',[
  (103,'Loser 101','Loser 102','Miami','2026-07-18','17:00'),
 ]),
 ('final',[
  (104,'Winner 101','Winner 102','New York / New Jersey','2026-07-19','15:00'),
 ]),
]

# ---- Construir lista de fixtures con numeraci├│n ----
fixtures=[]
n=0
for letter in ['A','B','C','D','E','F','G','H','I','J','K','L']:
    for (g,h,a,u,venue) in [m for m in gmatches if m[0]==letter]:
        n+=1
        fixtures.append({
            "id":f"m{n:03d}","no":n,"round":"grupos","group":g,
            "teamA":{"name":h[0],"flag":h[1]},"teamB":{"name":a[0],"flag":a[1]},
            "slotA":"","slotB":"",
            "kickoff":u.strftime('%Y-%m-%dT%H:%M:%SZ'),"kickoffMs":int(u.timestamp()*1000),
            "venue":venue,"editable":False
        })
for rkey,rows in KO:
    for (num,sa,sb,venue,iso,hhmm) in rows:
        d=dt.date.fromisoformat(iso); hh,mm=[int(x) for x in hhmm.split(':')]; u=et_to_utc(d,hh,mm)
        fixtures.append({
            "id":f"m{num:03d}","no":num,"round":rkey,"group":None,
            "teamA":{"name":"Por definir","flag":"ЁЯП│я╕П"},"teamB":{"name":"Por definir","flag":"ЁЯП│я╕П"},
            "slotA":slot(sa),"slotB":slot(sb),
            "kickoff":u.strftime('%Y-%m-%dT%H:%M:%SZ'),"kickoffMs":int(u.timestamp()*1000),
            "venue":venue,"editable":True
        })

assert len([f for f in fixtures if f['round']=='grupos'])==72, "grupos != 72"
assert len(fixtures)==104, f"total {len(fixtures)}"

# ---- Emitir tournament-data.js ----
def jflag(t): return f'["{t[0]}", "{t[1]}"]'
groups_js="{\n"
for letter in ['A','B','C','D','E','F','G','H','I','J','K','L']:
    groups_js+=f'  {letter}: [' + ", ".join(jflag(t) for t in groups[letter]) + "],\n"
groups_js+="}"

fx_js=json.dumps(fixtures, ensure_ascii=False, indent=2)
# json usa true/false/null -> v├бlidos en JS

out = '''/* =========================================================
 *  ESTRUCTURA DEL TORNEO тАУ MUNDIAL 2026 (datos oficiales)
 *  48 equipos en 12 grupos (AтАУL). Calendario real de fase de
 *  grupos y eliminatorias generado desde el fixture oficial.
 *  "kickoff" es un instante en UTC; la app lo muestra en la zona
 *  horaria local de cada usuario (Intl.DateTimeFormat).
 *  NO EDITAR A MANO: regenerar con _gen_fixtures.py si cambia.
 * ========================================================= */

// 12 grupos de 4 equipos [nombre, bandera]
const DEFAULT_GROUPS = %s;

// Rondas (orden y etiqueta) para filtros y bonus
const ROUNDS = [
  { key: "grupos",        label: "Grupos" },
  { key: "dieciseisavos", label: "16avos" },
  { key: "octavos",       label: "Octavos" },
  { key: "cuartos",       label: "Cuartos" },
  { key: "semis",         label: "Semis" },
  { key: "tercer",        label: "3er lugar" },
  { key: "final",         label: "Final" }
];

// Calendario oficial (104 partidos). Para eliminatorias, teamA/teamB
// son "Por definir" y slotA/slotB describen qui├йn clasifica
// (ej. "1┬║ Grupo A", "Mejor 3┬║", "Ganador #74"); el admin fija los
// equipos reales conforme avanza el torneo.
const WC_FIXTURES = %s;

/** Genera el mapa inicial de partidos {id: match} para un torneo nuevo. */
function buildInitialMatches() {
  const matches = {};
  WC_FIXTURES.forEach((f) => {
    matches[f.id] = Object.assign({}, f, {
      teamA: { name: f.teamA.name, flag: f.teamA.flag },
      teamB: { name: f.teamB.name, flag: f.teamB.flag },
      realA: null, realB: null, played: false
    });
  });
  return matches;
}

// Lista de equipos (para el selector de campe├│n)
function allTeamNames() {
  const set = [];
  Object.values(DEFAULT_GROUPS).forEach((g) => g.forEach((t) => set.push(t)));
  return set.sort((a, b) => a[0].localeCompare(b[0]));
}
''' % (groups_js, fx_js)

with open('js/tournament-data.js','w',encoding='utf-8') as f:
    f.write(out)
print("OK", len(fixtures), "fixtures; grupos:", {k:len(v) for k,v in groups.items()})
